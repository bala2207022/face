# index.py – Web dashboard for face-based class attendance

import os
import re
import json
import time
import base64
from datetime import datetime
from pathlib import Path
from typing import Dict

import numpy as np
import cv2
from flask import Flask, request, jsonify, render_template_string
try:
  from insightface.app import FaceAnalysis
  HAS_INSIGHTFACE = True
except Exception:
  FaceAnalysis = None
  HAS_INSIGHTFACE = False
  print("Warning: insightface not available; face recognition features are disabled.")
try:
  from openpyxl import Workbook, load_workbook
  from openpyxl.styles import Font, PatternFill, Alignment
  HAS_OPENPYXL = True
except Exception:
  Workbook = None
  load_workbook = None
  Font = PatternFill = Alignment = None
  HAS_OPENPYXL = False
  print("Warning: openpyxl not installed; Excel report features are disabled.")


# basic paths and configuration (relative to project root for cross-platform compatibility)
ROOT = os.path.normpath(os.path.join(os.path.dirname(__file__), ".."))
DATA_DIR = os.path.join(ROOT, "data", "students")
MODELS_DIR = os.path.join(ROOT, "models")
CENTROIDS = os.path.join(MODELS_DIR, "centroids.json")

THRESHOLD = float(os.getenv("THRESHOLD", "0.45"))
DET_SIZE = (640, 640)
COOLDOWN_S = 600  # minimum seconds between repeated logs for same face

Path(DATA_DIR).mkdir(parents=True, exist_ok=True)
Path(MODELS_DIR).mkdir(parents=True, exist_ok=True)


# Excel-based storage configuration
REPORTS_DIR = os.path.join(ROOT, "excel_reports")
CLASSES_META = os.path.join(MODELS_DIR, "classes_meta.json")
STUDENTS_META = os.path.join(MODELS_DIR, "students_meta.json")

Path(REPORTS_DIR).mkdir(parents=True, exist_ok=True)
Path(MODELS_DIR).mkdir(parents=True, exist_ok=True)


def load_json_file(path, default):
  if not os.path.exists(path):
    return default
  with open(path, "r") as f:
    try:
      return json.load(f)
    except Exception:
      return default


def save_json_file(path, data):
  with open(path, "w") as f:
    json.dump(data, f, indent=2, default=str)


def load_classes_meta():
  return load_json_file(CLASSES_META, {"next_id": 1, "classes": {}})


def save_classes_meta(meta):
  save_json_file(CLASSES_META, meta)


def load_students_meta():
  return load_json_file(STUDENTS_META, {"next_id": 1, "students": {}})


def save_students_meta(meta):
  save_json_file(STUDENTS_META, meta)


def _class_filepath_for_id(class_id: int) -> str | None:
  meta = load_classes_meta()
  cls = meta.get("classes", {}).get(str(class_id))
  if not cls:
    return None
  return cls.get("file")


def ensure_reports_setup():
  # ensure meta files exist
  m = load_classes_meta()
  save_classes_meta(m)
  s = load_students_meta()
  save_students_meta(s)


# student helpers (Excel-backed)
def upsert_student(face_label: str, student_name: str, student_code: str) -> int:
  meta = load_students_meta()
  students = meta.setdefault("students", {})
  if face_label in students:
    sid = students[face_label]["id"]
    students[face_label]["name"] = student_name
    students[face_label]["code"] = student_code
  else:
    sid = meta["next_id"]
    meta["next_id"] = sid + 1
    students[face_label] = {"id": sid, "name": student_name, "code": student_code}
  save_students_meta(meta)
  return sid


def get_student_id_by_face_label(face_label: str) -> int | None:
  meta = load_students_meta()
  s = meta.get("students", {}).get(face_label)
  if s:
    return s.get("id")
  return None


# professor and class helpers
def create_class(face_label: str, professor_name: str, professor_code: str, class_name: str) -> int:
  meta = load_classes_meta()
  cid = meta["next_id"]
  meta["next_id"] = cid + 1

  safe_name = re.sub(r"[^a-zA-Z0-9_-]", "_", class_name)
  filename = f"class_{cid}_{safe_name}.xlsx"
  filepath = os.path.join(REPORTS_DIR, filename)

  wb = Workbook()
  ws_meta = wb.active
  ws_meta.title = "meta"
  headers = [
    "class_id",
    "class_name",
    "professor_label",
    "professor_name",
    "professor_code",
    "start_time",
    "session_count",
  ]
  values = [
    cid,
    class_name,
    face_label,
    professor_name,
    professor_code,
    datetime.now().isoformat(),
    0,
  ]
  ws_meta.append(headers)
  ws_meta.append(values)

  ws_students = wb.create_sheet("students")
  ws_students.append(["student_id", "face_label", "student_name", "student_code", "total_present"])

  ws_sessions = wb.create_sheet("sessions")
  ws_sessions.append(["session_id", "timestamp", "student_id", "face_label", "student_name", "student_code"])

  wb.save(filepath)

  meta.setdefault("classes", {})[str(cid)] = {
    "file": filepath,
    "class_name": class_name,
    "professor_label": face_label,
    "professor_name": professor_name,
    "professor_code": professor_code,
    "start_time": datetime.now().isoformat(),
    "session_count": 0,
  }
  save_classes_meta(meta)
  return cid


def get_latest_class_for_professor(face_label: str):
  meta = load_classes_meta()
  best = None
  best_time = None
  for cid, info in meta.get("classes", {}).items():
    if info.get("professor_label") != face_label:
      continue
    st = info.get("start_time")
    if st is None:
      continue
    if best is None or st > best_time:
      best = cid
      best_time = st
  if best is None:
    return None, None
  return int(best), meta["classes"][best].get("class_name")


def mark_student_attendance(face_label: str, class_id: int):
  # Load class metadata and filepath
  meta = load_classes_meta()
  cls = meta.get("classes", {}).get(str(class_id))
  if not cls:
    return "Class", "Unknown", "Unknown"
  class_name = cls.get("class_name", "Class")
  filepath = cls.get("file")
  if not filepath or not os.path.exists(filepath):
    return class_name, "Unknown", "Unknown"

  # ensure student exists globally
  sid = get_student_id_by_face_label(face_label)
  if sid is None:
    sid = upsert_student(face_label, *parse_label(face_label))

  # current session id
  session_id = int(cls.get("session_count", 0))

  wb = load_workbook(filepath)
  ws_students = wb["students"]
  ws_sessions = wb["sessions"]

  # find or add student row
  student_row_idx = None
  for idx, row in enumerate(ws_students.iter_rows(min_row=2, values_only=False), start=2):
    cell_face = row[1].value
    if cell_face == face_label:
      student_row_idx = idx
      break
  if student_row_idx is None:
    # add new student
    name = parse_label(face_label)[1]
    code = parse_label(face_label)[0]
    ws_students.append([sid, face_label, name, code, 0])
    student_row_idx = ws_students.max_row

  # Prevent duplicate check-ins on the same calendar date
  today = datetime.now().date()
  already_today = False
  for r in ws_sessions.iter_rows(min_row=2, values_only=True):
    sess_id = r[0]
    ts = r[1]
    stuid = r[2]
    try:
      sess_date = datetime.fromisoformat(str(ts)).date()
    except Exception:
      sess_date = None
    if stuid == sid and sess_date == today:
      already_today = True
      break

  if already_today:
    # nothing to change
    student_name = parse_label(face_label)[1]
    student_code = parse_label(face_label)[0]
    return class_name, student_name, student_code, False

  # append session attendance (require an open session)
  if session_id > 0:
    now_ts = datetime.now().isoformat()
    ws_sessions.append([session_id, now_ts, sid, face_label, parse_label(face_label)[1], parse_label(face_label)[0]])
    # increment student's total_present
    total_cell = ws_students.cell(row=student_row_idx, column=5)
    try:
      total_cell.value = int(total_cell.value or 0) + 1
    except Exception:
      total_cell.value = 1

  wb.save(filepath)
  student_name = parse_label(face_label)[1]
  student_code = parse_label(face_label)[0]
  return class_name, student_name, student_code, True


def get_class_summary(class_id: int):
  filepath = _class_filepath_for_id(class_id)
  if not filepath or not os.path.exists(filepath):
    return 0, 0, 0, []

  meta = load_classes_meta()
  cls = meta.get("classes", {}).get(str(class_id), {})
  session_id = cls.get("session_count", 0)

  wb = load_workbook(filepath, data_only=True)
  ws_students = wb["students"]
  ws_sessions = wb["sessions"]

  total_students = max(0, ws_students.max_row - 1)

  present_ids = set()
  present_list = []
  if session_id > 0:
    for r in ws_sessions.iter_rows(min_row=2, values_only=True):
      if r[0] == session_id:
        present_ids.add(r[2])
        present_list.append((r[4], r[5]))

  total_present = len(present_ids)
  total_absent = max(0, total_students - total_present)
  return total_students, total_present, total_absent, present_list


def write_summary_sheet_for_class(class_id: int):
  """Compute per-student totals across all sessions and write a clear Summary sheet.

  Columns: name, id, date, no_of_present, no_of_absent, total_class
  """
  filepath = _class_filepath_for_id(class_id)
  if not filepath or not os.path.exists(filepath):
    return False

  meta = load_classes_meta()
  cls = meta.get("classes", {}).get(str(class_id), {})
  total_sessions = int(cls.get("session_count", 0))

  wb = load_workbook(filepath)
  ws_students = wb["students"]
  ws_sessions = wb["sessions"]

  # build mapping student_id -> (name, code)
  students = {}
  for r in ws_students.iter_rows(min_row=2, values_only=True):
    sid, face, name, code, total_present = r[0], r[1], r[2], r[3], r[4] if len(r) > 4 else 0
    students[sid] = {"name": name, "code": code, "face": face, "present": 0}

  # count presents across sessions
  session_dates = {}  # session_id -> timestamp
  for r in ws_sessions.iter_rows(min_row=2, values_only=True):
    sid_sess = r[0]
    ts = r[1]
    stuid = r[2]
    session_dates[sid_sess] = ts
    if stuid in students:
      students[stuid]["present"] = students[stuid].get("present", 0) + 1

  # prepare or replace Summary sheet
  if "Summary" in wb.sheetnames:
    ws = wb["Summary"]
    wb.remove(ws)
  ws = wb.create_sheet("Summary")

  # add header row with formatting
  headers = ["name", "id", "date", "no_of_present", "no_of_absent", "total_class"]
  ws.append(headers)

  # format header row: bold white text on blue background
  header_fill = PatternFill(start_color="0070C0", end_color="0070C0", fill_type="solid")
  header_font = Font(bold=True, color="FFFFFF")
  for cell in ws[1]:
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal="center", vertical="center")

  # choose date: most recent session date if available, else today
  recent_date = None
  if session_dates:
    # find latest session id by max
    latest_sid = max(session_dates.keys())
    recent_date_raw = session_dates.get(latest_sid)
    try:
      recent_date = str(recent_date_raw).split("T")[0]
    except Exception:
      recent_date = str(recent_date_raw)
  else:
    recent_date = datetime.now().strftime("%Y-%m-%d")

  for sid, info in students.items():
    name = info.get("name") or ""
    code = info.get("code") or ""
    present = int(info.get("present", 0))
    absent = max(0, total_sessions - present)
    ws.append([name, code, recent_date, present, absent, total_sessions])

  # format data rows: center alignment for numbers and alternate row colors
  for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=6):
    for cell in row:
      cell.alignment = Alignment(horizontal="center", vertical="center")
      # alternate row colors for readability
      if row[0].row % 2 == 0:
        cell.fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")

  # auto-adjust column widths based on content
  from openpyxl.utils import get_column_letter
  for col_num, header in enumerate(headers, 1):
    max_length = len(header)
    col_letter = get_column_letter(col_num)
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=col_num, max_col=col_num):
      for cell in row:
        try:
          if cell.value:
            max_length = max(max_length, len(str(cell.value)))
        except Exception:
          pass
    adjusted_width = min(max_length + 2, 50)  # add 2 for padding, max 50
    ws.column_dimensions[col_letter].width = adjusted_width

  wb.save(filepath)
  return True


# face helpers
_face_app = None
_last_log_times: Dict[str, float] = {}


def face_app():
  """Return the insightface FaceAnalysis instance or raise if not available."""
  global _face_app
  if not HAS_INSIGHTFACE:
    raise RuntimeError(
      "insightface is not installed or failed to import. Install insightface or run in an environment where it's available."
    )
  if _face_app is None:
    _face_app = FaceAnalysis(name="buffalo_l")
    _face_app.prepare(ctx_id=0, det_size=DET_SIZE)
  return _face_app


def norm_cos(a: np.ndarray, b: np.ndarray) -> float:
    a = a / np.linalg.norm(a)
    b = b / np.linalg.norm(b)
    return float(np.dot(a, b))


def parse_label(label: str):
    return label.split("_", 1) if "_" in label else (label, label)


def data_url_to_bgr(data_url: str) -> np.ndarray | None:
    m = re.match(r"^data:image/(png|jpeg);base64,(.+)$", data_url or "")
    if not m:
        return None
    raw = base64.b64decode(m.group(2))
    nparr = np.frombuffer(raw, np.uint8)
    return cv2.imdecode(nparr, cv2.IMREAD_COLOR)


def load_centroids() -> Dict[str, np.ndarray]:
    if not os.path.exists(CENTROIDS):
        return {}
    with open(CENTROIDS, "r") as f:
        data = json.load(f).get("centroids", {})
    return {k: np.array(v, dtype=np.float32) for k, v in data.items()}


def save_centroids(cents: Dict[str, np.ndarray]):
    with open(CENTROIDS, "w") as f:
        json.dump({"centroids": {k: v.tolist() for k, v in cents.items()}}, f)


def compute_centroid_for_folder(folder: str) -> np.ndarray | None:
  # If insightface is available use it to extract per-face normalized embeddings.
  # Otherwise fall back to a simple image feature so data flows (create/register) work.
  embs = []
  for p in Path(folder).glob("*"):
    if not p.is_file():
      continue
    img = cv2.imread(str(p))
    if img is None:
      continue
    if HAS_INSIGHTFACE:
      try:
        app = face_app()
        faces = app.get(img)
      except Exception:
        faces = []
      if not faces:
        continue
      f = max(
        faces,
        key=lambda z: (z.bbox[2] - z.bbox[0]) * (z.bbox[3] - z.bbox[1]),
      )
      embs.append(f.normed_embedding.astype(np.float32))
    else:
      feat = compute_image_feature(img)
      if feat is not None:
        embs.append(feat.astype(np.float32))
  if not embs:
    return None
  return np.mean(embs, axis=0)


def compute_image_feature(img: np.ndarray) -> np.ndarray | None:
  """Fallback feature extractor: resize, grayscale, histogram/flattened normalized vector.
  This is NOT a face embedding — it's a coarse image descriptor used only as a fallback.
  """
  try:
    # convert to grayscale and resize to fixed size
    g = cv2.cvtColor(img, cv2.COLOR_BGR2GRAY)
    r = cv2.resize(g, (128, 128), interpolation=cv2.INTER_LINEAR)
    # normalize to [0,1]
    v = r.astype(np.float32).ravel() / 255.0
    # L2-normalize
    norm = np.linalg.norm(v)
    if norm <= 1e-6:
      return None
    return v / norm
  except Exception:
    return None


# flask application
app = Flask(__name__)
ensure_reports_setup()

INDEX_HTML = """
<!doctype html>
<html lang="en">
<head>
<meta charset="utf-8">
<title>Face Attendance System</title>
<meta name="viewport" content="width=device-width, initial-scale=1">
<link href="https://cdn.jsdelivr.net/npm/bootstrap@5.3.3/dist/css/bootstrap.min.css" rel="stylesheet">
<style>
  body { background:#0b1020; color:#e9ecf1; }
  .glass { background:rgba(255,255,255,0.06); border:1px solid rgba(255,255,255,0.12); border-radius:16px; backdrop-filter: blur(8px); }
  .title { font-weight:700; letter-spacing:.5px; }
  .btn-pill { border-radius:999px; padding:.6rem 1.2rem; }
  video { border-radius:12px; border:1px solid rgba(255,255,255,.15); }
  .tiny { font-size:.9rem; opacity:.85; }
  .tab-btn-active { background:#0d6efd; color:white; }
</style>
</head>
<body>
<div class="container py-4">
  <div class="d-flex align-items-center justify-content-between mb-3">
    <h1 class="title">Face Recognition Attendance</h1>
    <span class="tiny">Allow camera access for this page in the browser settings.</span>
  </div>

  <div class="mb-3 d-flex gap-2">
    <button id="tabCreate" class="btn btn-outline-light btn-pill tab-btn-active">Create Class</button>
    <button id="tabCheck" class="btn btn-outline-light btn-pill">Check-In</button>
  </div>

  <div class="row g-4">
    <!-- Camera block -->
    <div class="col-lg-6">
      <div class="p-3 glass">
        <h4 class="mb-3">Camera</h4>
        <video id="cam" width="560" height="420" autoplay playsinline></video>
        <canvas id="grab" width="560" height="420" hidden></canvas>
        <div class="mt-3 d-flex gap-2">
          <button id="startCam" class="btn btn-secondary btn-pill">Start Camera</button>
          <button id="capture" class="btn btn-outline-light btn-pill">Capture Frame</button>
          <button id="stopCam" class="btn btn-outline-warning btn-pill">Stop Camera</button>
        </div>
        <div id="lastShotWrap" class="mt-3" hidden>
          <div class="tiny mb-1">Last captured frame</div>
          <img id="lastShot" class="img-fluid rounded" />
        </div>
      </div>
    </div>

    <!-- Right side: two modes -->
    <div class="col-lg-6">
      <!-- Create Class section -->
      <div id="createSection" class="p-3 glass">
        <h4 class="mb-3">Create Class (Professor Only)</h4>
        <div class="row g-2">
          <div class="col-md-4">
            <label class="form-label">Professor Name</label>
            <input id="profNameCreate" class="form-control" />
          </div>
          <div class="col-md-4">
            <label class="form-label">Professor ID</label>
            <input id="profCodeCreate" class="form-control" />
          </div>
          <div class="col-md-4">
            <label class="form-label">Class Name / Subject</label>
            <input id="classNameCreate" class="form-control" placeholder="e.g. MSBA-700" />
          </div>
        </div>

        <div class="mt-3">
          <p class="tiny mb-1">Step 1: Stand in front of the camera and capture several frames.</p>
          <button id="saveProfFrame" class="btn btn-success btn-pill">Save Professor Frame</button>
        </div>

        <div class="mt-3">
          <p class="tiny mb-1">Step 2: When you have enough frames, create the class.</p>
          <button id="createClass" class="btn btn-primary btn-pill">Create Class</button>
        </div>

        <div id="createMsg" class="tiny mt-3"></div>
      </div>

      <!-- Check-In section -->
      <div id="checkSection" class="p-3 glass" style="display:none; margin-top:0;">
        <h4 class="mb-3">Check-In</h4>

        <h6>Step 1: Professor opens the class</h6>
        <p class="tiny">Professor stands in front of the camera and clicks the button below.</p>
        <button id="openClassBtn" class="btn btn-primary btn-pill mb-2">Scan Professor and Open Class</button>
        <div class="tiny mb-3">
          Active class id: <span id="activeClassId">None</span>,
          name: <span id="activeClassName">None</span>
        </div>

        <hr>

        <h6>New student registration for this class</h6>
        <p class="tiny">Use this only when a student is not registered. They will be linked to the active class.</p>
        <div class="row g-2">
          <div class="col-md-6">
            <label class="form-label">Student Name</label>
            <input id="studentName" class="form-control" />
          </div>
          <div class="col-md-6">
            <label class="form-label">Student ID</label>
            <input id="studentCode" class="form-control" />
          </div>
        </div>
        <div class="mt-2 d-flex gap-2">
          <button id="saveStudentFrame" class="btn btn-success btn-pill">Save Student Frame</button>
          <button id="registerStudent" class="btn btn-outline-light btn-pill">Register Student</button>
        </div>
        <div id="studentMsg" class="tiny mt-2"></div>

        <hr class="mt-4">

        <h6>Step 2: Students check in</h6>
        <p class="tiny">Students stand in front of the camera and click the button below.</p>
        <button id="studentCheckBtn" class="btn btn-info btn-pill">Student Check-In</button>
        <button id="summaryBtn" class="btn btn-outline-light btn-pill ms-2">Show Summary</button>

        <div class="mt-3">
          <div id="checkStatus" class="fs-5"></div>
          <div id="checkDetail" class="tiny"></div>
        </div>

        <div class="mt-3" id="summaryBox"></div>
      </div>
    </div>
  </div>
</div>

<script>
let stream = null;
let currentClassId = null;
let currentClassName = null;

const v  = document.getElementById('cam');
const c  = document.getElementById('grab');
const cx = c.getContext('2d');

function setTab(createActive){
  const createBtn = document.getElementById('tabCreate');
  const checkBtn  = document.getElementById('tabCheck');
  const createSec = document.getElementById('createSection');
  const checkSec  = document.getElementById('checkSection');
  if(createActive){
    createBtn.classList.add('tab-btn-active');
    checkBtn.classList.remove('tab-btn-active');
    createSec.style.display = '';
    checkSec.style.display  = 'none';
  }else{
    checkBtn.classList.add('tab-btn-active');
    createBtn.classList.remove('tab-btn-active');
    createSec.style.display = 'none';
    checkSec.style.display  = '';
  }
}

document.getElementById('tabCreate').onclick = ()=> setTab(true);
document.getElementById('tabCheck').onclick  = ()=> setTab(false);

async function startCam(){
  try{
    stream = await navigator.mediaDevices.getUserMedia({video:true, audio:false});
    v.srcObject = stream;
  }catch(e){
    alert('Camera blocked: ' + e.message);
  }
}
function stopCam(){
  if(stream){ stream.getTracks().forEach(t=>t.stop()); stream = null; }
}
function snap(){
  if(!v || !v.srcObject){ return null; }
  cx.drawImage(v, 0, 0, c.width, c.height);
  const dataUrl = c.toDataURL('image/jpeg', 0.95);
  const img = document.getElementById('lastShot');
  const wrap = document.getElementById('lastShotWrap');
  img.src = dataUrl;
  wrap.hidden = false;
  return dataUrl;
}

document.getElementById('startCam').onclick = startCam;
document.getElementById('stopCam').onclick  = stopCam;
document.getElementById('capture').onclick  = snap;

// professor frame saving during create class
document.getElementById('saveProfFrame').onclick = async ()=>{
  const profName  = document.getElementById('profNameCreate').value.trim();
  const profCode  = document.getElementById('profCodeCreate').value.trim();
  const className = document.getElementById('classNameCreate').value.trim();
  const dataUrl   = snap();
  if(!profName || !profCode || !className){
    alert('Fill professor name, id, and class name first.');
    return;
  }
  if(!dataUrl){
    alert('Start camera and capture a frame.');
    return;
  }
  const r = await fetch('/api/save_prof_frame', {
    method:'POST', headers:{'Content-Type':'application/json'},
    body: JSON.stringify({profName, profCode, className, dataUrl})
  });
  const j = await r.json();
  document.getElementById('createMsg').innerText = j.message || j.error || '';
};

// create class after capturing frames
document.getElementById('createClass').onclick = async ()=>{
  const profName  = document.getElementById('profNameCreate').value.trim();
  const profCode  = document.getElementById('profCodeCreate').value.trim();
  const className = document.getElementById('classNameCreate').value.trim();
  if(!profName || !profCode || !className){
    alert('Fill professor name, id, and class name first.');
    return;
  }
  const r = await fetch('/api/create_class', {
    method:'POST', headers:{'Content-Type':'application/json'},
    body: JSON.stringify({profName, profCode, className})
  });
  const j = await r.json();
  document.getElementById('createMsg').innerText = j.message || j.error || '';
};

// open class by professor scan
document.getElementById('openClassBtn').onclick = async ()=>{
  const dataUrl = snap();
  if(!dataUrl){
    alert('Start camera and capture a frame.');
    return;
  }
  const r = await fetch('/api/open_class', {
    method:'POST', headers:{'Content-Type':'application/json'},
    body: JSON.stringify({dataUrl})
  });
  const j = await r.json();
  if(!j.ok){
    document.getElementById('checkStatus').innerText = j.error || 'Could not open class.';
    document.getElementById('checkDetail').innerText = '';
    return;
  }
  currentClassId = j.classId;
  currentClassName = j.className;
  document.getElementById('activeClassId').innerText = j.classId;
  document.getElementById('activeClassName').innerText = j.className;
  document.getElementById('checkStatus').innerText = j.message || '';
  document.getElementById('checkDetail').innerText = '';
  document.getElementById('summaryBox').innerHTML = '';
  document.getElementById('studentMsg').innerText = '';
};

// save student frame for this class
document.getElementById('saveStudentFrame').onclick = async ()=>{
  if(!currentClassId){
    alert('Professor must open the class first.');
    return;
  }
  const studentName = document.getElementById('studentName').value.trim();
  const studentCode = document.getElementById('studentCode').value.trim();
  const dataUrl     = snap();
  if(!studentName || !studentCode){
    alert('Enter student name and id first.');
    return;
  }
  if(!dataUrl){
    alert('Start camera and capture a frame.');
    return;
  }
  const r = await fetch('/api/save_student_frame', {
    method:'POST', headers:{'Content-Type':'application/json'},
    body: JSON.stringify({studentName, studentCode, classId: currentClassId, dataUrl})
  });
  const j = await r.json();
  document.getElementById('studentMsg').innerText = j.message || j.error || '';
};

// register student (train and link to class)
document.getElementById('registerStudent').onclick = async ()=>{
  if(!currentClassId){
    alert('Professor must open the class first.');
    return;
  }
  const studentName = document.getElementById('studentName').value.trim();
  const studentCode = document.getElementById('studentCode').value.trim();
  if(!studentName || !studentCode){
    alert('Enter student name and id first.');
    return;
  }
  const r = await fetch('/api/register_student', {
    method:'POST', headers:{'Content-Type':'application/json'},
    body: JSON.stringify({studentName, studentCode, classId: currentClassId})
  });
  const j = await r.json();
  document.getElementById('studentMsg').innerText = j.message || j.error || '';
};

// student check-in
document.getElementById('studentCheckBtn').onclick = async ()=>{
  if(!currentClassId){
    alert('Professor must open the class first.');
    return;
  }
  const dataUrl = snap();
  if(!dataUrl){
    alert('Start camera and capture a frame.');
    return;
  }
  const r = await fetch('/api/check_student', {
    method:'POST', headers:{'Content-Type':'application/json'},
    body: JSON.stringify({dataUrl, classId: currentClassId})
  });
  const j = await r.json();
  document.getElementById('checkStatus').innerText = j.status || '';
  document.getElementById('checkDetail').innerText = j.detail || '';
};

// show summary
document.getElementById('summaryBtn').onclick = async ()=>{
  if(!currentClassId){
    alert('Professor must open the class first.');
    return;
  }
  const r = await fetch('/api/summary', {
    method:'POST', headers:{'Content-Type':'application/json'},
    body: JSON.stringify({classId: currentClassId})
  });
  const j = await r.json();
  if(!j.ok){
    alert(j.error || 'Could not load summary.');
    return;
  }
  let html = '';
  html += '<div class="tiny">Total students: ' + j.total_students +
          ' | Present: ' + j.total_present +
          ' | Absent: ' + j.total_absent + '</div>';
  if(j.present && j.present.length){
    html += '<div class="tiny mt-2">Present students:</div><ul class="tiny">';
    for(const s of j.present){
      html += '<li>' + s.name + ' (' + s.code + ')</li>';
    }
    html += '</ul>';
  }
  document.getElementById('summaryBox').innerHTML = html;
};

// try to start camera automatically
document.addEventListener('DOMContentLoaded', ()=>{ setTimeout(startCam, 500); });
</script>
</body>
</html>
"""


@app.get("/")
def index():
    return render_template_string(INDEX_HTML)


# save professor training frames
@app.post("/api/save_prof_frame")
def api_save_prof_frame():
    data = request.get_json(force=True)
    prof_name = (data.get("profName") or "").strip()
    prof_code = (data.get("profCode") or "").strip()
    class_name = (data.get("className") or "").strip()
    data_url = data.get("dataUrl") or ""

    if not prof_name or not prof_code or not class_name:
        return jsonify(ok=False, error="Missing professor or class information"), 400

    img = data_url_to_bgr(data_url)
    if img is None:
        return jsonify(ok=False, error="No image"), 400

    label = f"prof_{prof_code}_{prof_name}"
    folder = os.path.join(DATA_DIR, label)
    Path(folder).mkdir(parents=True, exist_ok=True)
    out = os.path.join(folder, f"{int(time.time() * 1000)}.jpg")
    cv2.imwrite(out, img)
    n = len(list(Path(folder).glob("*.jpg")))
    return jsonify(ok=True, message=f"Saved frame for professor. Total frames: {n}")


# create class after professor frames are captured
@app.post("/api/create_class")
def api_create_class():
    data = request.get_json(force=True)
    prof_name = (data.get("profName") or "").strip()
    prof_code = (data.get("profCode") or "").strip()
    class_name = (data.get("className") or "").strip()

    if not prof_name or not prof_code or not class_name:
        return jsonify(ok=False, error="Missing professor or class information"), 400

    if not HAS_OPENPYXL:
      return jsonify(ok=False, error="openpyxl is required for creating classes (install openpyxl in the active Python environment)."), 501

    label = f"prof_{prof_code}_{prof_name}"
    folder = os.path.join(DATA_DIR, label)
    if not os.path.isdir(folder):
        return jsonify(ok=False, error="No professor frames saved yet"), 400

    centroid = compute_centroid_for_folder(folder)
    if centroid is None:
        return jsonify(ok=False, error="No faces detected in saved frames"), 400

    cents = load_centroids()
    cents[label] = centroid
    save_centroids(cents)

    class_id = create_class(label, prof_name, prof_code, class_name)
    msg = f"Class created successfully. Class id = {class_id}, name = {class_name}"
    return jsonify(ok=True, classId=class_id, message=msg)


# save student frames for active class
@app.post("/api/save_student_frame")
def api_save_student_frame():
    data = request.get_json(force=True)
    student_name = (data.get("studentName") or "").strip()
    student_code = (data.get("studentCode") or "").strip()
    class_id = data.get("classId")
    data_url = data.get("dataUrl") or ""

    try:
        int(class_id)
    except (TypeError, ValueError):
        return jsonify(ok=False, error="Invalid class id"), 400

    if not student_name or not student_code:
        return jsonify(ok=False, error="Missing student information"), 400

    img = data_url_to_bgr(data_url)
    if img is None:
        return jsonify(ok=False, error="No image"), 400

    label = f"{student_code}_{student_name}"
    folder = os.path.join(DATA_DIR, label)
    Path(folder).mkdir(parents=True, exist_ok=True)
    out = os.path.join(folder, f"{int(time.time() * 1000)}.jpg")
    cv2.imwrite(out, img)
    n = len(list(Path(folder).glob("*.jpg")))
    return jsonify(ok=True, message=f"Saved frame for student. Total frames: {n}")


# register student for active class (train centroid and link to class)
@app.post("/api/register_student")
def api_register_student():
    data = request.get_json(force=True)
    student_name = (data.get("studentName") or "").strip()
    student_code = (data.get("studentCode") or "").strip()
    class_id = data.get("classId")

    try:
        class_id = int(class_id)
    except (TypeError, ValueError):
        return jsonify(ok=False, error="Invalid class id"), 400

    if not student_name or not student_code:
        return jsonify(ok=False, error="Missing student information"), 400

    if not HAS_OPENPYXL:
      return jsonify(ok=False, error="openpyxl is required for registering students (install openpyxl in the active Python environment)."), 501

    label = f"{student_code}_{student_name}"
    folder = os.path.join(DATA_DIR, label)
    if not os.path.isdir(folder):
        return jsonify(ok=False, error="No student frames saved yet"), 400

    centroid = compute_centroid_for_folder(folder)
    if centroid is None:
        return jsonify(ok=False, error="No faces detected in saved frames"), 400

    cents = load_centroids()
    cents[label] = centroid
    save_centroids(cents)

    upsert_student(label, student_name, student_code)
    student_id = get_student_id_by_face_label(label)
    if student_id is None:
      return jsonify(ok=False, error="Could not resolve student id after insert"), 500

    # link student to class workbook
    filepath = _class_filepath_for_id(class_id)
    if not filepath or not os.path.exists(filepath):
      return jsonify(ok=False, error="Class workbook not found"), 500

    wb = load_workbook(filepath)
    ws_students = wb["students"]

    exists = False
    for r in ws_students.iter_rows(min_row=2, values_only=True):
      if r[1] == label or r[0] == student_id:
        exists = True
        break
    if not exists:
      ws_students.append([student_id, label, student_name, student_code, 0])
      wb.save(filepath)

    msg = f"Student registered and linked to class {class_id}."
    return jsonify(ok=True, message=msg)


# open class by scanning professor face
@app.post("/api/open_class")
def api_open_class():
    data = request.get_json(force=True)
    data_url = data.get("dataUrl") or ""

    img = data_url_to_bgr(data_url)
    if img is None:
        return jsonify(ok=False, error="No image"), 400

    if not HAS_OPENPYXL:
      return jsonify(ok=False, error="openpyxl is required to open classes and record sessions (install openpyxl in the active Python environment)."), 501

    cents = load_centroids()
    if not cents:
      return jsonify(ok=False, error="Model not trained yet"), 200

    # Extract embedding: prefer insightface face embedding, otherwise use a fallback image feature
    if HAS_INSIGHTFACE:
      appface = face_app()
      faces = appface.get(img)
      if not faces:
        return jsonify(ok=False, error="No face detected"), 200
      f = max(
          faces,
          key=lambda z: (z.bbox[2] - z.bbox[0]) * (z.bbox[3] - z.bbox[1]),
      )
      emb = f.normed_embedding.astype(np.float32)
    else:
      emb = compute_image_feature(img)
      if emb is None:
        return jsonify(ok=False, error="Could not compute fallback image feature"), 200

    best_lab, best_sim = "Unknown", -1.0
    for lab, c in cents.items():
        s = norm_cos(emb, c)
        if s > best_sim:
            best_sim, best_lab = s, lab

    if best_sim < THRESHOLD or not best_lab.startswith("prof_"):
        return jsonify(ok=False, error="Professor not recognized"), 200

    class_id, class_name = get_latest_class_for_professor(best_lab)
    if class_id is None:
      return jsonify(ok=False, error="No class found for this professor"), 200

    # increment session count for this class and create a session entry
    meta = load_classes_meta()
    cls = meta.get("classes", {}).get(str(class_id))
    if cls is None:
      return jsonify(ok=False, error="Class metadata missing"), 200
    cls["session_count"] = int(cls.get("session_count", 0)) + 1
    session_id = cls["session_count"]
    save_classes_meta(meta)

    # append session row in workbook
    filepath = cls.get("file")
    if filepath and os.path.exists(filepath):
      wb = load_workbook(filepath)
      ws_sessions = wb["sessions"]
      ws_sessions.append([session_id, datetime.now().isoformat(), "", "", "", ""])
      wb.save(filepath)

    msg = f"Welcome Professor. Class opened: {class_name} (id {class_id})"
    return jsonify(ok=True, classId=class_id, className=class_name, message=msg)


# student check-in for an open class
@app.post("/api/check_student")
def api_check_student():
    data = request.get_json(force=True)
    data_url = data.get("dataUrl") or ""
    class_id = data.get("classId")

    try:
        class_id = int(class_id)
    except (TypeError, ValueError):
        return jsonify(ok=False, error="Invalid class id"), 400

    img = data_url_to_bgr(data_url)
    if img is None:
        return jsonify(ok=False, error="No image"), 400

    cents = load_centroids()
    if not cents:
      return jsonify(ok=False, error="Model not trained yet"), 200

    if HAS_INSIGHTFACE:
      appface = face_app()
      faces = appface.get(img)
      if not faces:
        return jsonify(ok=True, status="No face detected", detail=""), 200
      f = max(
        faces,
        key=lambda z: (z.bbox[2] - z.bbox[0]) * (z.bbox[3] - z.bbox[1]),
      )
      emb = f.normed_embedding.astype(np.float32)
    else:
      emb = compute_image_feature(img)
      if emb is None:
        return jsonify(ok=True, status="No face detected", detail=""), 200

    best_lab, best_sim = "Unknown", -1.0
    for lab, c in cents.items():
        s = norm_cos(emb, c)
        if s > best_sim:
            best_sim, best_lab = s, lab

    ts = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    if best_sim < THRESHOLD:
        return jsonify(
            ok=True,
            status="Not registered",
            detail=f"time={ts}  sim={best_sim:.2f}",
        )

    if best_lab.startswith("prof_"):
        return jsonify(
            ok=True,
            status="Professor detected",
            detail=f"time={ts}  sim={best_sim:.2f}",
        )

    now = time.time()
    if now - _last_log_times.get(best_lab, 0) <= COOLDOWN_S:
        status = "Already marked present for this class"
        detail = f"time={ts}  sim={best_sim:.2f}"
        return jsonify(ok=True, status=status, detail=detail)

    class_name, student_name, student_code, recorded = mark_student_attendance(
      best_lab, class_id
    )
    _last_log_times[best_lab] = now

    if not recorded:
      status = "Already marked present today"
      detail = f"{student_name} ({student_code})  |  time={ts}  sim={best_sim:.2f}"
      return jsonify(ok=True, status=status, detail=detail)

    status = f"Welcome to {class_name}"
    detail = f"{student_name} ({student_code})  |  time={ts}  sim={best_sim:.2f}"
    return jsonify(ok=True, status=status, detail=detail)


# class summary endpoint
@app.post("/api/summary")
def api_summary():
    data = request.get_json(force=True)
    class_id = data.get("classId")
    try:
        class_id = int(class_id)
    except (TypeError, ValueError):
        return jsonify(ok=False, error="Invalid class id"), 400
    total_students, total_present, total_absent, present_list = get_class_summary(class_id)

    # write a clear Summary sheet in the workbook with requested columns
    wrote = write_summary_sheet_for_class(class_id)

    present = [{"name": name, "code": code} for (name, code) in present_list]
    return jsonify(
        ok=True,
        total_students=total_students,
        total_present=total_present,
        total_absent=total_absent,
        present=present,
        summary_written=wrote,
    )


if __name__ == "__main__":
    # run with: python src/dashboard.py
    app.run(host="127.0.0.1", port=50135, debug=True)