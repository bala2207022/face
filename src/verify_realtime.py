import os
import re
import json
import time
from datetime import datetime
from pathlib import Path

import numpy as np
import cv2
from insightface.app import FaceAnalysis
from openpyxl import Workbook, load_workbook

# paths and basic configuration (project-relative, universal)
ROOT = os.path.normpath(os.path.join(os.path.dirname(__file__), ".."))
MODELS_DIR = os.path.join(ROOT, "models")
REPORTS_DIR = os.path.join(ROOT, "excel_reports")
MODEL_FILE = os.path.join(MODELS_DIR, "centroids.json")

THRESHOLD = 0.45          # higher -> more strict match
DET_SIZE = (640, 640)
LOG_COOLDOWN_SEC = 600    # do not mark same face again within 10 minutes

Path(REPORTS_DIR).mkdir(parents=True, exist_ok=True)

# state for current session
MODE = "PROFESSOR"  # first recognized face is professor
CLASS_INFO = {
    "class_name": None,
    "prof_name": None,
    "prof_code": None,
    "file_path": None,
}
STUDENT_META = {}          # face_label -> (student_name, student_code)
SESSION_ATTENDANCE = []    # list of dicts with attendance entries


def cos_sim(a: np.ndarray, b: np.ndarray) -> float:
    a = a / np.linalg.norm(a)
    b = b / np.linalg.norm(b)
    return float(np.dot(a, b))


def open_cam():
    """Try a few indices/backends so it works on most systems."""
    tries = [
        (0, cv2.CAP_AVFOUNDATION),
        (0, cv2.CAP_QT),
        (0, cv2.CAP_ANY),
        (1, cv2.CAP_ANY),
    ]
    for idx, backend in tries:
        cap = cv2.VideoCapture(idx, backend)
        if cap.isOpened():
            print(f"Opened camera index={idx}, backend={backend}")
            return cap
        cap.release()
    print("Could not open any camera. Check permissions / close other apps.")
    return None


def load_centroids():
    if not os.path.exists(MODEL_FILE):
        print(f"Centroids file not found: {MODEL_FILE}")
        print("Run build_embeddings.py and train_centroid.py first.")
        return {}
    with open(MODEL_FILE, "r") as f:
        raw = json.load(f).get("centroids", {})
    return {k: np.array(v, dtype=np.float32) for k, v in raw.items()}


def create_class_excel(prof_face_label: str):
    """Ask professor details and create an Excel file for this class."""
    print("\nProfessor detected. Please enter class details.")
    prof_name = input("Professor Name: ").strip()
    prof_code = input("Professor ID: ").strip()
    class_name = input("Class Name / Subject: ").strip()

    if not class_name:
        class_name = "Class"
    safe_class = re.sub(r"[^A-Za-z0-9_-]+", "_", class_name)
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"class_{safe_class}_{ts}.xlsx"
    file_path = os.path.join(REPORTS_DIR, filename)

    wb = Workbook()
    ws = wb.active
    ws.title = "Attendance"
    ws.append(
        [
            "Class Name",
            "Professor Name",
            "Professor ID",
            "Face Label",
            "Student ID",
            "Student Name",
            "Date",
            "Time",
        ]
    )
    wb.save(file_path)

    CLASS_INFO["class_name"] = class_name
    CLASS_INFO["prof_name"] = prof_name
    CLASS_INFO["prof_code"] = prof_code
    CLASS_INFO["file_path"] = file_path

    print(f"\nClass created and Excel file started:")
    print(f"  Class Name : {class_name}")
    print(f"  Professor  : {prof_name} ({prof_code})")
    print(f"  File       : {file_path}\n")
    print("Students can start checking in now.\n")


def append_attendance_to_excel(face_label: str, student_name: str, student_code: str, dt: datetime):
    """Append one attendance row to the Excel sheet."""
    file_path = CLASS_INFO["file_path"]
    if not file_path:
        return

    wb = load_workbook(file_path)
    ws = wb["Attendance"]

    ws.append(
        [
            CLASS_INFO["class_name"],
            CLASS_INFO["prof_name"],
            CLASS_INFO["prof_code"],
            face_label,
            student_code,
            student_name,
            dt.strftime("%Y-%m-%d"),
            dt.strftime("%H:%M:%S"),
        ]
    )
    wb.save(file_path)


def mark_student_attendance(face_label: str, frame) -> None:
    """Handle student attendance: ask name/id if first time, log to Excel."""
    if face_label in STUDENT_META:
        student_name, student_code = STUDENT_META[face_label]
    else:
        print(f"\nNew student detected for label: {face_label}")
        student_name = input("Student Name: ").strip()
        student_code = input("Student ID: ").strip()
        STUDENT_META[face_label] = (student_name, student_code)

    dt = datetime.now()

    # Prevent duplicate check-ins on the same calendar date
    file_path = CLASS_INFO.get("file_path")
    already_today = False
    if file_path and os.path.exists(file_path):
        wb = load_workbook(file_path, data_only=True)
        if "Attendance" in wb.sheetnames:
            ws = wb["Attendance"]
            for r in ws.iter_rows(min_row=2, values_only=True):
                try:
                    row_face = r[3]
                    row_date = r[6]
                except Exception:
                    row_face = None
                    row_date = None
                if row_face == face_label and row_date == dt.strftime("%Y-%m-%d"):
                    already_today = True
                    break

    if already_today:
        cv2.putText(
            frame,
            "Already checked in today",
            (20, 30),
            cv2.FONT_HERSHEY_SIMPLEX,
            0.7,
            (0, 200, 255),
            2,
        )
        print(f"Already checked in today: {student_name} ({student_code})")
        return

    append_attendance_to_excel(face_label, student_name, student_code, dt)

    SESSION_ATTENDANCE.append(
        {
            "face_label": face_label,
            "student_name": student_name,
            "student_code": student_code,
            "datetime": dt,
        }
    )

    cv2.putText(
        frame,
        f"Welcome to the class {CLASS_INFO['class_name']}",
        (20, 30),
        cv2.FONT_HERSHEY_SIMPLEX,
        0.7,
        (0, 255, 0),
        2,
    )
    cv2.putText(
        frame,
        f"{student_name} ({student_code})",
        (20, 60),
        cv2.FONT_HERSHEY_SIMPLEX,
        0.7,
        (0, 255, 0),
        2,
    )

    print(f"Marked present: {student_name} ({student_code})")


def show_class_summary():
    """Print a summary and write a summary sheet into the Excel file."""
    if not CLASS_INFO["file_path"]:
        return

    if not SESSION_ATTENDANCE:
        print("\nNo students checked in. No summary to write.")
        return

    # distinct students by student_code
    summary_map = {}
    for rec in SESSION_ATTENDANCE:
        code = rec["student_code"]
        name = rec["student_name"]
        if code not in summary_map:
            summary_map[code] = {"name": name, "count": 0}
        summary_map[code]["count"] += 1

    # compute total classes by counting unique dates in the Attendance sheet
    wb = load_workbook(CLASS_INFO["file_path"], data_only=True)
    ws_att = wb["Attendance"]
    dates = set()
    for r in ws_att.iter_rows(min_row=2, values_only=True):
        try:
            d = r[6]
        except Exception:
            d = None
        if d:
            dates.add(str(d))
    total_classes = max(1, len(dates))

    total_present = len(summary_map)

    print("\nAttendance summary for this session")
    print(f"Class Name : {CLASS_INFO['class_name']}")
    print(f"Professor  : {CLASS_INFO['prof_name']} ({CLASS_INFO['prof_code']})")
    print(f"Total present students: {total_present}\n")
    print("Present students:")
    for code, info in summary_map.items():
        print(f"- {info['name']} ({code})  times_present={info['count']}")

    # write summary sheet in Excel with requested columns
    if "Summary" in wb.sheetnames:
        ws = wb["Summary"]
        wb.remove(ws)
    ws = wb.create_sheet("Summary")
    headers = ["name", "id", "date", "no_of_present", "no_of_absent", "total_class"]
    ws.append(headers)

    # format header row: bold white text on blue background
    from openpyxl.styles import Font, PatternFill, Alignment
    header_fill = PatternFill(start_color="0070C0", end_color="0070C0", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # date for this summary: use most recent attendance date or today
    recent_date = None
    if dates:
        recent_date = sorted(dates)[-1].split("T")[0]
    else:
        recent_date = datetime.now().strftime("%Y-%m-%d")

    for code, info in summary_map.items():
        present = info.get("count", 0)
        absent = max(0, total_classes - present)
        ws.append([info.get("name"), code, recent_date, present, absent, total_classes])

    # format data rows: center alignment and alternate row colors
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=6):
        for cell in row:
            cell.alignment = Alignment(horizontal="center", vertical="center")
            # alternate row colors for readability
            if row[0].row % 2 == 0:
                cell.fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")

    # auto-adjust column widths based on content
    from openpyxl.utils import get_column_letter
    for col_num, header in enumerate(headers, 1):
        max_length = len(header)
        col_letter = get_column_letter(col_num)
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=col_num, max_col=col_num):
            for cell in row:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except Exception:
                    pass
        adjusted_width = min(max_length + 2, 50)  # add 2 for padding, max 50
        ws.column_dimensions[col_letter].width = adjusted_width

    wb.save(CLASS_INFO["file_path"])
    print(f"\nSummary written to: {CLASS_INFO['file_path']}")


def main():
    global MODE

    print("Face Recognition Attendance System (Excel mode)")
    print("First recognized face will be the professor to start the class.")
    print("Then students can check in.\n")

    centroids = load_centroids()
    if not centroids:
        return

    app = FaceAnalysis(name="buffalo_l")
    app.prepare(ctx_id=0, det_size=DET_SIZE)

    cap = open_cam()
    if cap is None:
        return

    print("Press 'q' to quit.\n")
    last_logged = {}  # face_label -> last log timestamp

    while True:
        ret, frame = cap.read()
        if not ret or frame is None:
            dbg = np.zeros((360, 640, 3), dtype=np.uint8)
            cv2.putText(
                dbg,
                "No frame from camera. Check permissions / close other apps.",
                (12, 180),
                cv2.FONT_HERSHEY_SIMPLEX,
                0.55,
                (0, 0, 255),
                2,
            )
            cv2.imshow("Verification", dbg)
            if cv2.waitKey(30) & 0xFF == ord("q"):
                break
            continue

        faces = app.get(frame)

        if not faces:
            cv2.putText(
                frame,
                "No face detected",
                (20, 30),
                cv2.FONT_HERSHEY_SIMPLEX,
                0.8,
                (0, 200, 255),
                2,
            )
        else:
            f = max(
                faces,
                key=lambda z: (z.bbox[2] - z.bbox[0]) * (z.bbox[3] - z.bbox[1]),
            )
            emb = f.normed_embedding

            best_lab, best_sim = "Unknown", -1.0
            for lab, c in centroids.items():
                s = cos_sim(emb, c)
                if s > best_sim:
                    best_sim, best_lab = s, lab

            recognized = best_sim >= THRESHOLD
            name_to_draw = best_lab if recognized else "Not Registered"

            x1, y1, x2, y2 = map(int, f.bbox)
            color = (0, 200, 0) if recognized else (0, 0, 255)
            cv2.rectangle(frame, (x1, y1), (x2, y2), color, 2)
            cv2.putText(
                frame,
                f"{name_to_draw}  sim={best_sim:.2f}",
                (x1, max(20, y1 - 10)),
                cv2.FONT_HERSHEY_SIMPLEX,
                0.7,
                color,
                2,
            )

            if recognized:
                face_label = best_lab
                now = time.time()

                if MODE == "PROFESSOR" and CLASS_INFO["file_path"] is None:
                    # first recognized face starts the class and Excel
                    create_class_excel(face_label)
                    MODE = "STUDENT"
                    cv2.putText(
                        frame,
                        "Class started",
                        (20, 100),
                        cv2.FONT_HERSHEY_SIMPLEX,
                        0.8,
                        (255, 255, 0),
                        2,
                    )
                elif MODE == "STUDENT" and CLASS_INFO["file_path"] is not None:
                    # debounce repeated logs
                    if now - last_logged.get(face_label, 0) > LOG_COOLDOWN_SEC:
                        mark_student_attendance(face_label, frame)
                        last_logged[face_label] = now

        cv2.imshow("Verification", frame)
        if cv2.waitKey(1) & 0xFF == ord("q"):
            break

    cap.release()
    cv2.destroyAllWindows()

    # write summary when session ends
    show_class_summary()


if __name__ == "__main__":
    main()
