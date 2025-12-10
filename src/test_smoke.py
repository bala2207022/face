#!/usr/bin/env python3
"""Smoke-test: create a sample class Excel workbook and exercise duplicate check logic.

This script does not depend on insightface or OpenCV — it only requires openpyxl.
Run with the project's venv activated:

  python src/test_smoke.py

It writes `excel_reports/smoke_test_class.xlsx` and attempts to mark the same student twice
to demonstrate duplicate-prevention logic.
"""
from pathlib import Path
from datetime import date
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

ROOT = Path(__file__).resolve().parents[1]
OUT = ROOT / "excel_reports"
OUT.mkdir(exist_ok=True)

FILE = OUT / "smoke_test_class.xlsx"

wb = Workbook()

# meta sheet
meta = wb.active
meta.title = "meta"
meta["A1"] = "class_id"
meta["B1"] = "smoke_test_class"

# students
students = wb.create_sheet("students")
students.append(["name", "student_id"])
students.append(["Test Student", "S001"])

# Attendance
att = wb.create_sheet("Attendance")
att.append(["student_id", "name", "date"])

def mark_attendance(student_id: str, name: str):
    today = date.today().isoformat()
    # duplicate check
    for row in att.iter_rows(min_row=2, values_only=True):
        if row[0] == student_id and row[2] == today:
            print(f"Duplicate detected for {student_id} on {today}")
            return False
    att.append([student_id, name, today])
    print(f"Marked present: {student_id} - {name} on {today}")
    return True

mark_attendance("S001", "Test Student")
mark_attendance("S001", "Test Student")  # should be duplicate

# Summary sheet
summary = wb.create_sheet("Summary")
header = ["name", "student_id", "date", "no_of_present", "no_of_absent", "total_class"]
summary.append(header)
for c in range(1, len(header) + 1):
    cell = summary.cell(row=1, column=c)
    cell.font = Font(bold=True)
    cell.fill = PatternFill(start_color="FFBDE9F8", end_color="FFBDE9F8", fill_type="solid")
    cell.alignment = Alignment(horizontal="center")

summary.append(["Test Student", "S001", date.today().isoformat(), 1, 0, 1])

wb.save(FILE)
print(f"Smoke test workbook written to: {FILE}")
